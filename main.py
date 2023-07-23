import os
from typing import List, Dict

import fitz
import openpyxl
from openpyxl.styles import PatternFill

FONT_SIZE_PARAGRAPH = 9.5
FONT_SIZE_TITLE = 9
FONT_SIZE_AUTHORS = 9
FONT_SIZE_AFFILIATIONS = 8
FONT_SIZE_ABSTRACT = 9.134002685546875

HEADER_BACKGROUND_COLOR = "749BFF"


def create_colored_header_style(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def extract_information_from_pdf(file_path: str, start_page: int) -> List[Dict]:
    """
    Extracts information from a PDF file starting from the specified page.

    Args:
        file_path (str): The path to the PDF file.
        start_page (int): The page number to start extracting information from.

    Returns:
        List[Dict]: A list of dictionaries containing the extracted information.
    """
    with fitz.open(file_path) as file:
        previous_article = None

        for page_number in range(start_page - 1, file.page_count):
            page = file.load_page(page_number)
            text_blocks = page.get_text("dict")["blocks"]

            for article in text_blocks:
                current_article = {
                    "authors": [],
                    "affiliations": [],
                    "session": "",
                    "title": "",
                    "abstract": "",
                }

                for line in article["lines"]:
                    for span in line["spans"]:
                        is_paragraph = span["font"] == "TimesNewRomanPS-BoldItal" and span[
                            "size"] == FONT_SIZE_PARAGRAPH
                        is_title = span["font"] == "TimesNewRomanPS-BoldMT" and span["size"] == FONT_SIZE_TITLE
                        is_authors = span["font"] == "TimesNewRomanPS-ItalicMT" and span["size"] == FONT_SIZE_AUTHORS
                        is_affiliations = span["font"] == "TimesNewRomanPS-ItalicMT" and span[
                            "size"] == FONT_SIZE_AFFILIATIONS
                        is_abstract = span["size"] == FONT_SIZE_ABSTRACT

                        text_pattern = span["text"]

                        if is_paragraph:
                            current_article["session"] += text_pattern
                        elif is_title:
                            current_article["title"] += text_pattern
                        elif is_authors:
                            current_article["authors"].append(text_pattern)
                        elif is_affiliations:
                            current_article["affiliations"].append(text_pattern)
                        elif is_abstract:
                            current_article["abstract"] += text_pattern

                if article["type"] == 0:
                    if current_article["session"] or current_article["title"] or current_article["authors"] or \
                            current_article["affiliations"] or current_article["abstract"]:
                        if not current_article["session"]:
                            if previous_article is not None:
                                previous_article["title"] += current_article["title"]
                                previous_article["affiliations"].extend(current_article["affiliations"])
                                previous_article["abstract"] += current_article["abstract"]
                                current_article = previous_article
                        yield current_article

                previous_article = current_article


def merge_information_blocks(blocks: List[Dict]) -> List[Dict]:
    """
    Merges information blocks from sessions into a single block.

    Args:
        blocks (List[Dict]): A list of dictionaries containing information about extracted blocks.

    Returns:
        List[Dict]: A list of dictionaries containing the merged information.
    """
    merged_blocks = []
    current_block = None

    for block in blocks:
        if not current_block:
            current_block = block
        elif block["session"]:
            merged_blocks.append(current_block)
            current_block = block
        else:
            current_block["title"] += block["title"]
            current_block["affiliations"].extend(block["affiliations"])
            current_block["abstract"] += block["abstract"]

    if current_block:
        merged_blocks.append(current_block)

    return merged_blocks


def save_to_excel_file(blocks: List[Dict], file_path: str) -> None:
    """
    Saves information to an Excel file.

    Args:
        blocks (List[Dict]): A list of dictionaries containing information to be saved.
        file_path (str): The path to the Excel file.
    """
    if not os.path.exists(file_path):
        create_excel_file_with_headers(file_path)

    existing_authors = set()

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for block in blocks:
        session = block["session"]
        title = block["title"]
        authors = block["authors"]
        affiliations = block["affiliations"]
        abstract = block["abstract"]

        if session and tuple(authors) in existing_authors:
            continue

        for author in authors:
            if author.startswith(","):
                author = author[2:]
            author_tuple = (session, author)
            if author_tuple in existing_authors:
                continue
            existing_authors.add(author_tuple)

            row = sheet.max_row + 1
            sheet.cell(row=row, column=1).value = author
            sheet.cell(row=row, column=2).value = ", ".join(affiliations)
            sheet.cell(row=row, column=3).value = session
            sheet.cell(row=row, column=4).value = ""
            sheet.cell(row=row, column=5).value = title
            sheet.cell(row=row, column=6).value = abstract

    workbook.save(file_path)


def create_excel_file_with_headers(file_path: str) -> None:
    """
    Creates an Excel file with column headers.

    Args:
        file_path (str): The path to the Excel file.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name (incl, titles if any mentioned"
    sheet["B1"] = "Affiliations"
    sheet["C1"] = "Session name"
    sheet["D1"] = "Persons Location"
    sheet["E1"] = "Topic Title"
    sheet["F1"] = "Presentation Abstract"
    workbook.save(file_path)


if __name__ == "__main__":
    pdf_file = "book.pdf"
    excel_file = "result.xlsx"
    start_page = 44

    pdf_information = extract_information_from_pdf(pdf_file, start_page)
    merged_information_blocks = merge_information_blocks(pdf_information)
    save_to_excel_file(merged_information_blocks, excel_file)

    if not os.path.exists(excel_file):
        create_excel_file_with_headers(excel_file)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    for col in range(1, 7):
        cell = sheet.cell(row=1, column=col)
        cell.fill = create_colored_header_style(HEADER_BACKGROUND_COLOR)

    workbook.save(excel_file)

    save_to_excel_file(merged_information_blocks, excel_file)
